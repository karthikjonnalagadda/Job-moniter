"""Excel export engine (openpyxl).

Produces a multi-sheet, professionally formatted workbook:

    Summary · Top Matches · All Jobs · Duplicate Jobs · Company Statistics ·
    Skill Gap Analysis · Collector Statistics · Pipeline Metrics · Search History

Features: styled frozen headers, auto column width, auto-filters, Excel tables,
conditional formatting (match-score colour scale), apply-link hyperlinks, and
bar charts on the statistics sheets.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.exporters.report_base import ReportExporter
from app.exporters.rows import COLUMNS, job_to_row
from app.models.report_record import ReportFormat

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from app.reports.dataset import ReportData

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True, color="1F4E78")
_LINK_FONT = Font(color="0563C1", underline="single")
_MAX_WIDTH = 60


class ExcelExporter(ReportExporter):
    """Render a ``ReportData`` to a formatted ``.xlsx`` workbook."""

    format = ReportFormat.EXCEL
    extension = "xlsx"

    def export(self, data: ReportData, destination: Path) -> Path:
        wb = Workbook()
        wb.remove(wb.active)  # drop the default sheet
        self._summary_sheet(wb, data)
        self._jobs_sheet(wb, "Top Matches", data.top_matches, apply_links=True, color_scale=True)
        self._jobs_sheet(wb, "All Jobs", data.jobs, apply_links=True, color_scale=True)
        self._duplicates_sheet(wb, data)
        self._counts_sheet(wb, "Company Statistics", "Company", data.analytics.companies)
        self._counts_sheet(wb, "Skill Gap Analysis", "Missing Skill", data.skill_gap)
        self._collector_sheet(wb, data)
        self._pipeline_sheet(wb, data)
        self._history_sheet(wb, data)
        wb.save(destination)
        return destination

    # ---- sheets ---------------------------------------------------------
    def _summary_sheet(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Summary")
        ws["A1"] = data.title
        ws["A1"].font = _TITLE_FONT
        a = data.analytics
        rows = [
            ("Generated", data.generated_at.isoformat() if data.generated_at else ""),
            ("Run ID", data.run_id or ""),
            ("Total jobs", a.total_jobs),
            ("Ranked jobs", a.ranked_jobs),
            ("New today", data.new_today),
            ("Average match", a.average_match),
            ("Duplicate groups", len(data.duplicate_groups)),
            ("Report version", data.versions.report_version),
            ("Schema version", data.versions.schema_version),
            ("Pipeline version", data.versions.pipeline_version),
        ]
        for i, (label, value) in enumerate(rows, start=3):
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = Font(bold=True)
            ws[f"B{i}"] = value
        self._autowidth(ws)

    def _jobs_sheet(
        self, wb: Workbook, title: str, jobs: list[Any], *, apply_links: bool, color_scale: bool
    ) -> None:
        ws = wb.create_sheet(title)
        self._write_header(ws, COLUMNS)
        for job in jobs:
            row = job_to_row(job)
            ws.append([row.get(col) for col in COLUMNS])
        self._finalise_table(ws, title.replace(" ", ""), len(COLUMNS), len(jobs))

        if apply_links:
            self._hyperlink_column(ws, "apply_url", len(jobs))
        if color_scale and jobs:
            self._color_scale(ws, "match_score", len(jobs))

    def _duplicates_sheet(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Duplicate Jobs")
        headers = ("Fingerprint", "Count", "Postings")
        self._write_header(ws, headers)
        for group in data.duplicate_groups:
            ws.append([group.fingerprint[:16], group.count, " | ".join(group.labels)])
        self._finalise_table(ws, "Duplicates", len(headers), len(data.duplicate_groups))

    def _counts_sheet(self, wb: Workbook, title: str, label: str, stats: list[Any]) -> None:
        ws = wb.create_sheet(title)
        self._write_header(ws, (label, "Count"))
        for stat in stats:
            ws.append([stat.label, stat.count])
        self._finalise_table(ws, title.replace(" ", ""), 2, len(stats))
        if stats:
            self._bar_chart(ws, title, rows=len(stats))

    def _collector_sheet(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Collector Statistics")
        headers = ("Collector", "Runs", "Jobs Found", "Duplicates", "Errors", "Avg Response ms")
        self._write_header(ws, headers)
        for c in data.collector_stats:
            avg = c.total_response_ms / c.response_samples if c.response_samples else 0.0
            ws.append([c.collector, c.runs, c.total_jobs_found, c.total_duplicates,
                       c.total_errors, round(avg, 1)])
        self._finalise_table(ws, "Collectors", len(headers), len(data.collector_stats))

    def _pipeline_sheet(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Pipeline Metrics")
        headers = ("Run ID", "Collected", "Normalized", "Filtered Out", "Duplicates",
                   "Stored", "Duration s")
        self._write_header(ws, headers)
        for run in data.pipeline_runs:
            ws.append([run.run_id[:12], run.collected, run.normalized, run.filtered_out,
                       run.duplicates, run.stored, run.duration_seconds])
        self._finalise_table(ws, "Pipeline", len(headers), len(data.pipeline_runs))

    def _history_sheet(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Search History")
        headers = ("Run ID", "Started", "Status", "Collected", "Stored")
        self._write_header(ws, headers)
        for run in data.pipeline_runs:
            ws.append([run.run_id[:12], run.started_at.isoformat() if run.started_at else "",
                       str(run.status), run.collected, run.stored])
        self._finalise_table(ws, "History", len(headers), len(data.pipeline_runs))

    # ---- styling helpers ------------------------------------------------
    def _write_header(self, ws: Worksheet, headers: tuple[str, ...]) -> None:
        ws.append(list(headers))
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    def _finalise_table(self, ws: Worksheet, name: str, n_cols: int, n_rows: int) -> None:
        last_col = get_column_letter(n_cols)
        ref = f"A1:{last_col}{n_rows + 1}"
        ws.auto_filter.ref = ref
        if n_rows > 0:  # openpyxl tables require at least one data row
            table = Table(displayName=f"tbl{name}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(table)
        self._autowidth(ws)

    def _autowidth(self, ws: Worksheet) -> None:
        for column_cells in ws.columns:
            lengths = (len(str(c.value)) for c in column_cells if c.value is not None)
            length = max(lengths, default=8)
            letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[letter].width = min(_MAX_WIDTH, max(10, length + 2))

    def _hyperlink_column(self, ws: Worksheet, column_name: str, n_rows: int) -> None:
        col_idx = COLUMNS.index(column_name) + 1
        for r in range(2, n_rows + 2):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.font = _LINK_FONT

    def _color_scale(self, ws: Worksheet, column_name: str, n_rows: int) -> None:
        col = get_column_letter(COLUMNS.index(column_name) + 1)
        rule = ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=70, mid_color="FFEB84",
            end_type="num", end_value=100, end_color="63BE7B",
        )
        ws.conditional_formatting.add(f"{col}2:{col}{n_rows + 1}", rule)

    def _bar_chart(self, ws: Worksheet, title: str, *, rows: int) -> None:
        chart = BarChart()
        chart.title = title
        chart.type = "bar"
        chart.height = 8
        chart.width = 16
        data = Reference(ws, min_col=2, min_row=1, max_row=min(rows + 1, 16))
        cats = Reference(ws, min_col=1, min_row=2, max_row=min(rows + 1, 16))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")
