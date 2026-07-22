"""Exporters: CSV (plain/compressed/streaming), JSON, Excel, HTML, PDF."""

from __future__ import annotations

import gzip
import json
from datetime import UTC, datetime
from pathlib import Path

import pytest
from app.exporters.csv_exporter import CsvExporter
from app.exporters.excel import ExcelExporter
from app.exporters.html import HtmlExporter
from app.exporters.json_exporter import JsonExporter
from app.exporters.pdf import PdfExporter
from app.exporters.registry import build_exporter
from app.models.analytics import AnalyticsReport, CountStat
from app.models.common import Location, SalaryRange
from app.models.job import Job, MatchDetail
from app.models.report_record import ReportFormat
from app.reports.dataset import DuplicateGroup, ReportData


def _job(i: int, company: str, score: float) -> Job:
    return Job(
        job_hash=f"h{i}", external_id=str(i), source="greenhouse", company_name=company,
        role="ML Engineer", normalized_role="ML Engineer", url=f"https://x/{i}",
        location=Location(city="Bangalore", country="IN"),
        salary=SalaryRange(
            min_amount=2_500_000, max_amount=3_500_000, currency="INR", period="year"
        ),
        skills=["Python", "FastAPI"], technologies=["Python"],
        posted_date=datetime(2026, 7, 21, tzinfo=UTC),
        match=MatchDetail(score=score, similarity=0.9, skill=1.0, missing_skills=["Docker"]),
    )


def _data() -> ReportData:
    jobs = [_job(1, "Google", 88), _job(2, "Flipkart", 72), _job(3, "Zomato", 45)]
    return ReportData(
        generated_at=datetime(2026, 7, 21, 12, 0, tzinfo=UTC),
        run_id="run123",
        analytics=AnalyticsReport(
            total_jobs=3, ranked_jobs=3, average_match=68.3,
            companies=[CountStat(label="Google", count=1), CountStat(label="Flipkart", count=1)],
            roles=[CountStat(label="ML Engineer", count=3)],
            skills=[CountStat(label="Python", count=3)],
        ),
        jobs=jobs, top_matches=jobs,
        skill_gap=[CountStat(label="Docker", count=3)],
        duplicate_groups=[DuplicateGroup(fingerprint="fp", count=2, labels=["A", "B"])],
    )


def test_csv_plain_and_compressed(tmp_path: Path) -> None:
    out = CsvExporter().export(_data(), tmp_path / "r.csv")
    text = out.read_text(encoding="utf-8")
    assert "company" in text.splitlines()[0]
    assert "Google" in text

    gz = CsvExporter(compress=True).export(_data(), tmp_path / "r.csv.gz")
    with gzip.open(gz, "rt", encoding="utf-8") as fh:
        assert "Google" in fh.read()


def test_json_pretty_compact_streaming(tmp_path: Path) -> None:
    pretty = JsonExporter(pretty=True).export(_data(), tmp_path / "p.json")
    payload = json.loads(pretty.read_bytes())
    assert payload["meta"]["run_id"] == "run123"
    assert len(payload["jobs"]) == 3

    compact = JsonExporter(pretty=False).export(_data(), tmp_path / "c.json")
    assert b"\n  " not in compact.read_bytes()  # no indentation

    nd = JsonExporter(streaming=True).export(_data(), tmp_path / "s.ndjson")
    lines = nd.read_bytes().splitlines()
    assert len(lines) == 3 and json.loads(lines[0])["company"] == "Google"


def test_excel_sheets_and_features(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    out = ExcelExporter().export(_data(), tmp_path / "r.xlsx")
    wb = load_workbook(out)
    assert wb.sheetnames == [
        "Summary", "Top Matches", "All Jobs", "Duplicate Jobs", "Company Statistics",
        "Skill Gap Analysis", "Collector Statistics", "Pipeline Metrics", "Search History",
    ]
    ws = wb["Top Matches"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert any(c.hyperlink for row in ws.iter_rows(min_row=2) for c in row)


def test_html_report_and_theme() -> None:
    html = HtmlExporter().render(_data())
    assert "<title>" in html
    assert 'href="https://x/1"' in html
    assert "bar-fill" in html  # CSS chart
    assert "#0f1720" in HtmlExporter(theme="dark").render(_data())


def test_pdf_report(tmp_path: Path) -> None:
    pytest.importorskip("reportlab")
    out = PdfExporter().export(_data(), tmp_path / "r.pdf")
    assert out.read_bytes()[:4] == b"%PDF"


def test_registry_resolves_formats() -> None:
    for fmt in ReportFormat:
        assert build_exporter(fmt).format == fmt
